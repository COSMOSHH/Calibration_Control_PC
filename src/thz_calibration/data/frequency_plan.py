from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class SignalSourceFrequencies:
    """Frequency pair derived from the IF/LO calculation workbook."""

    calibration_frequency_ghz: float
    if_frequency_ghz: float
    lo_frequency_ghz: float


class FrequencyPlan:
    """Lookup IF and LO source frequencies for a target THz output frequency."""

    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = Path(workbook_path)

    def lookup(self, calibration_frequency_ghz: float) -> SignalSourceFrequencies:
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"frequency plan workbook not found: {self.workbook_path}")

        values_wb = load_workbook(self.workbook_path, data_only=True, read_only=True)
        formula_wb = load_workbook(self.workbook_path, data_only=False, read_only=True)
        try:
            values_ws = values_wb.active
            formula_ws = formula_wb[values_ws.title]
            target = float(calibration_frequency_ghz)
            for row in range(2, values_ws.max_row + 1):
                target_frequency = self._as_number(values_ws.cell(row=row, column=3).value)
                if target_frequency is None or abs(target_frequency - target) > 1e-6:
                    continue

                if_frequency = self._as_number(values_ws.cell(row=row, column=1).value)
                lo_frequency = self._as_number(values_ws.cell(row=row, column=2).value)
                if lo_frequency is None:
                    lo_frequency = self._calculate_lo_frequency(formula_ws, row, target_frequency, if_frequency)
                if if_frequency is None or lo_frequency is None:
                    raise ValueError(f"frequency plan row {row} is missing IF/LO values")

                return SignalSourceFrequencies(
                    calibration_frequency_ghz=target_frequency,
                    if_frequency_ghz=if_frequency,
                    lo_frequency_ghz=lo_frequency,
                )
        finally:
            values_wb.close()
            formula_wb.close()

        raise ValueError(f"calibration frequency {calibration_frequency_ghz:g} GHz is not in the frequency plan")

    def _calculate_lo_frequency(
        self,
        formula_ws,
        row: int,
        target_frequency_ghz: float,
        if_frequency_ghz: float | None,
    ) -> float | None:
        """Fallback for workbooks whose formula cells have no cached value."""
        if if_frequency_ghz is None:
            return None

        formula = formula_ws.cell(row=row, column=2).value
        if isinstance(formula, str) and formula.strip().replace(" ", "").upper() == f"=(C{row}-A{row})/12":
            return (target_frequency_ghz - if_frequency_ghz) / 12.0
        return None

    def _as_number(self, value) -> float | None:
        if isinstance(value, bool) or value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                return float(text)
            except ValueError:
                return None
        return None
