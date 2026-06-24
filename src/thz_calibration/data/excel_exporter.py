from __future__ import annotations

from pathlib import Path
from shutil import copy2

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from ..config import OUTPUT_DIR, TEMPLATE_DIR
from ..models import CalibrationResult
from ..utils import format_phase


RED_FONT = Font(color="FF0000")

MULTI_BEAM_RESULT_NAME = "CalData_MultiFeed_MultiBeamDir_212to224.xlsx"
MULTI_BEAM_FREQUENCIES_GHZ = (212, 215, 218, 221, 224)
MULTI_BEAM_HEADER_ROW = 2
MULTI_BEAM_DATA_START_ROW = 3
MULTI_BEAM_BEAM_COL = 1
MULTI_BEAM_MEASURED_PHASE_COLS = (6, 7, 8, 9)


class ExcelExporter:
    """Read and write UI0 calibration workbooks."""

    def __init__(self, template_dir: Path = TEMPLATE_DIR, output_dir: Path = OUTPUT_DIR) -> None:
        self.template_dir = Path(template_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def save_scan_result(self, result: CalibrationResult) -> Path:
        """Save one feed scan result in the existing process workbook format."""
        if result.config.target_feed_id == 1:
            return self._save_feed1_test_data(result)
        return self._save_cal_process(result)

    def save_multi_feed_result(self, frequency_ghz: float, beam_angle_deg: float) -> Path:
        """Generate or update the multi-frequency, multi-beam summary workbook."""
        return self._save_multi_beam_result(frequency_ghz, beam_angle_deg)

    def _save_feed1_test_data(self, result: CalibrationResult) -> Path:
        """Save Feed1 single-feed power test data."""
        output = self._copy_or_create("TestData_Feed1_BDp30_212.xlsx", "TestData")
        wb = load_workbook(output)
        ws = wb.active
        ws["A1"] = result.config.frequency_ghz

        best = result.best_point
        self._clear_rows(ws, start_row=3, max_col=2)
        for row_index, point in enumerate(result.points, start=3):
            ws.cell(row=row_index, column=1, value=point.phase_deg)
            ws.cell(row=row_index, column=2, value=round(point.average_power_uw, 6))
            if point is best:
                self._mark_cells_red(ws, row_index, (1, 2))

        wb.save(output)
        return output

    def _save_cal_process(self, result: CalibrationResult) -> Path:
        """Save Feed2~4 phase scan process data."""
        ref = "".join(str(feed_id) for feed_id in range(1, result.config.target_feed_id))
        output_name = f"CalProcess_Feed{result.config.target_feed_id}wrt{ref}_BDp30_212.xlsx"
        output = self._copy_or_create(output_name, "CalProcess")

        wb = load_workbook(output)
        ws = wb.active
        ws["A1"] = result.config.frequency_ghz
        ws["B1"] = result.config.beam_angle_deg

        self._write_reference_rows(ws, result.config.target_feed_id)

        data_start_row = result.config.target_feed_id * 2 + 1
        best = result.best_point
        self._clear_data_columns(ws, start_row=data_start_row, columns=(2, 3))
        for row_index, point in enumerate(result.points, start=data_start_row):
            ws.cell(row=row_index, column=2, value=point.phase_deg)
            ws.cell(row=row_index, column=3, value=round(point.average_power_uw, 6))
            if point is best:
                self._mark_cells_red(ws, row_index, (2, 3))

        wb.save(output)
        return output

    def read_best_feed_point(self, feed_id: int) -> tuple[float, float] | None:
        """Return (phase_deg, power_uw) for the highest-power point of a feed."""
        workbook = self._output_path_for_feed(feed_id)
        if workbook is None or not workbook.exists():
            return None

        wb = load_workbook(workbook, data_only=True)
        ws = wb.active
        start_row = 3 if feed_id == 1 else feed_id * 2 + 1
        return self._read_best_point(ws, start_row)

    def _write_reference_rows(self, ws: Worksheet, target_feed_id: int) -> None:
        """Write previous feeds' best points into the current process workbook."""
        for feed_id in range(1, target_feed_id):
            best = self.read_best_feed_point(feed_id)
            if best is None:
                continue
            phase_deg, power_uw = best
            row = feed_id * 2 + 1
            ws.cell(row=row, column=2, value=phase_deg)
            ws.cell(row=row, column=3, value=round(power_uw, 6))
            self._mark_cells_red(ws, row, (2, 3))

    def _read_best_point(self, ws: Worksheet, start_row: int) -> tuple[float, float] | None:
        """Scan down from start_row and select the row with maximum power."""
        best: tuple[float, float] | None = None
        for row in range(start_row, ws.max_row + 1):
            phase = ws.cell(row=row, column=2 if start_row > 3 else 1).value
            power = ws.cell(row=row, column=3 if start_row > 3 else 2).value
            if not isinstance(phase, (int, float)) or not isinstance(power, (int, float)):
                continue
            if best is None or power > best[1]:
                best = (float(phase), float(power))
        return best

    def _output_path_for_feed(self, feed_id: int) -> Path | None:
        """Locate the process workbook that stores a feed's best point."""
        if feed_id == 1:
            return self.output_dir / "TestData_Feed1_BDp30_212.xlsx"
        if 2 <= feed_id <= 4:
            ref = "".join(str(current) for current in range(1, feed_id))
            return self.output_dir / f"CalProcess_Feed{feed_id}wrt{ref}_BDp30_212.xlsx"
        return None

    def _save_multi_beam_result(self, frequency_ghz: float, beam_angle_deg: float) -> Path:
        """Write Feed1~4 apparent phases to F:I in the new summary workbook."""
        missing: list[int] = []
        measured_phases: list[float] = []
        for feed_id in range(1, 5):
            best = self.read_best_feed_point(feed_id)
            if best is None:
                missing.append(feed_id)
                continue
            measured_phases.append(best[0])

        if missing:
            missing_text = ", ".join(f"Feed{feed_id}" for feed_id in missing)
            raise ValueError(f"Missing best point data required for final summary: {missing_text}")

        output, created = self._copy_or_create_multi_beam_result()
        wb = load_workbook(output)
        self._ensure_multi_beam_frequency_sheets(wb)
        if created:
            self._clear_multi_beam_measurements_from_workbook(wb)

        ws = self._multi_beam_sheet_for_frequency(wb, frequency_ghz)
        row = self._find_multi_beam_row(ws, beam_angle_deg)
        if row is None:
            if created:
                wb.save(output)
            raise ValueError(
                f"{frequency_ghz:g} GHz sheet does not contain Beam Dir. "
                f"{beam_angle_deg:g} deg in column A. Add that row in {MULTI_BEAM_RESULT_NAME} first."
            )

        for phase_deg, column in zip(measured_phases, MULTI_BEAM_MEASURED_PHASE_COLS):
            ws.cell(row=row, column=column, value=round(phase_deg, 6))

        wb.save(output)
        return output

    def _copy_or_create_multi_beam_result(self) -> tuple[Path, bool]:
        """Create the cumulative multi-beam workbook only when it is missing."""
        output = self.output_dir / MULTI_BEAM_RESULT_NAME
        if output.exists():
            return output, False

        template = self._template_path(MULTI_BEAM_RESULT_NAME)
        if template.exists():
            copy2(template, output)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = str(MULTI_BEAM_FREQUENCIES_GHZ[0])
            self._initialize_multi_beam_sheet(ws, MULTI_BEAM_FREQUENCIES_GHZ[0])
            wb.save(output)
        return output, True

    def _ensure_multi_beam_frequency_sheets(self, wb: Workbook) -> None:
        """Ensure only the required frequency worksheets exist and use the new layout."""
        layout_ws = self._find_multi_beam_layout_sheet(wb)
        for frequency in MULTI_BEAM_FREQUENCIES_GHZ:
            title = str(frequency)
            ws = wb[title] if title in wb.sheetnames else None
            if ws is not None and self._is_multi_beam_sheet(ws):
                ws["B1"] = frequency
                self._refresh_multi_beam_theory_formulas(ws)
                continue

            if ws is not None and layout_ws is None:
                self._initialize_multi_beam_sheet(ws, frequency)
                self._refresh_multi_beam_theory_formulas(ws)
                layout_ws = ws
                continue

            if ws is not None:
                wb.remove(ws)

            if layout_ws is not None:
                ws = wb.copy_worksheet(layout_ws)
                ws.title = title
                self._clear_multi_beam_measurements(ws)
            else:
                ws = wb.create_sheet(title)
                self._initialize_multi_beam_sheet(ws, frequency)
                layout_ws = ws
            ws["B1"] = frequency
            self._refresh_multi_beam_theory_formulas(ws)

        expected_titles = {str(freq) for freq in MULTI_BEAM_FREQUENCIES_GHZ}
        for ws in list(wb.worksheets):
            if ws.title in expected_titles:
                continue
            if ws.title.isdigit() and 212 <= int(ws.title) <= 224:
                wb.remove(ws)
                continue
            if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
                wb.remove(ws)
        self._order_multi_beam_frequency_sheets(wb)

    def _order_multi_beam_frequency_sheets(self, wb: Workbook) -> None:
        """Keep frequency worksheets in the required output order."""
        ordered_titles = [str(freq) for freq in MULTI_BEAM_FREQUENCIES_GHZ]
        by_title = {ws.title: ws for ws in wb.worksheets}
        ordered = [by_title[title] for title in ordered_titles if title in by_title]
        extras = [ws for ws in wb.worksheets if ws.title not in set(ordered_titles)]
        wb._sheets = ordered + extras

    def _find_multi_beam_layout_sheet(self, wb: Workbook) -> Worksheet | None:
        """Find a worksheet that already matches the new summary layout."""
        for ws in wb.worksheets:
            if self._is_multi_beam_sheet(ws):
                return ws
        return None

    def _is_multi_beam_sheet(self, ws: Worksheet) -> bool:
        """Detect the new summary table by the stable Beam Dir. header."""
        header = ws.cell(row=MULTI_BEAM_HEADER_ROW, column=MULTI_BEAM_BEAM_COL).value
        return (
            isinstance(header, str)
            and header.strip().lower().startswith("beam")
            and ws.max_column >= max(MULTI_BEAM_MEASURED_PHASE_COLS)
        )

    def _initialize_multi_beam_sheet(self, ws: Worksheet, frequency: int) -> None:
        """Create a minimal fallback sheet if the Excel template is unavailable."""
        for merged_range in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(merged_range))
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).value = None

        ws["A1"] = "Freq."
        ws["B1"] = frequency
        ws.merge_cells("C1:E1")
        ws["C1"] = "theoretical phase"
        ws.merge_cells("G1:I1")
        ws["G1"] = "apparent phase measured value"
        headers = [
            "Beam Dir.",
            "phi1(theoretical value)",
            "phi2(theoretical value)",
            "phi3(theoretical value)",
            "phi4(theoretical value)",
            "phip1/deg",
            "phip2/deg",
            "phip3/deg",
            "phip4/deg",
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=MULTI_BEAM_HEADER_ROW, column=col, value=header)

        for row, beam_angle in enumerate(range(-30, 31), start=MULTI_BEAM_DATA_START_ROW):
            ws.cell(row=row, column=MULTI_BEAM_BEAM_COL, value=beam_angle)
        self._refresh_multi_beam_theory_formulas(ws)

    def _refresh_multi_beam_theory_formulas(self, ws: Worksheet) -> None:
        """Set B:E theory formulas on every prepared Beam Dir. row."""
        for row in range(MULTI_BEAM_DATA_START_ROW, ws.max_row + 1):
            beam_angle = self._as_number(ws.cell(row=row, column=MULTI_BEAM_BEAM_COL).value)
            if beam_angle is None:
                continue
            ws.cell(row=row, column=2, value=0)
            for offset, column in enumerate((3, 4, 5), start=1):
                ws.cell(
                    row=row,
                    column=column,
                    value=f"=MOD(-180*$B$1*{offset}*15.52*2*SIN(A{row}*PI()/180)/300,360)",
                )

    def _clear_multi_beam_measurements_from_workbook(self, wb: Workbook) -> None:
        """Remove example measured values copied from the template workbook."""
        for ws in wb.worksheets:
            if self._is_multi_beam_sheet(ws):
                self._clear_multi_beam_measurements(ws)

    def _clear_multi_beam_measurements(self, ws: Worksheet) -> None:
        """Clear F:I measured phase columns while preserving Beam Dir. and theory."""
        for row in range(MULTI_BEAM_DATA_START_ROW, ws.max_row + 1):
            for column in MULTI_BEAM_MEASURED_PHASE_COLS:
                ws.cell(row=row, column=column).value = None

    def _multi_beam_sheet_for_frequency(self, wb: Workbook, frequency_ghz: float) -> Worksheet:
        """Return the worksheet for one of the required GHz frequencies."""
        title = self._multi_beam_sheet_title(frequency_ghz)
        return wb[title]

    def _multi_beam_sheet_title(self, frequency_ghz: float) -> str:
        """Map an allowed GHz frequency to its worksheet title."""
        frequency = float(frequency_ghz)
        frequency_int = round(frequency)
        if abs(frequency - frequency_int) > 1e-6:
            raise ValueError("Final summary frequency must be one of: 212, 215, 218, 221, 224 GHz.")
        if frequency_int not in MULTI_BEAM_FREQUENCIES_GHZ:
            raise ValueError("Final summary frequency must be one of: 212, 215, 218, 221, 224 GHz.")
        return str(frequency_int)

    def _find_multi_beam_row(self, ws: Worksheet, beam_angle_deg: float) -> int | None:
        """Find the manually prepared Beam Dir. row for the current measurement."""
        target = float(beam_angle_deg)
        for row in range(MULTI_BEAM_DATA_START_ROW, ws.max_row + 1):
            value = self._as_number(ws.cell(row=row, column=MULTI_BEAM_BEAM_COL).value)
            if value is not None and abs(value - target) <= 1e-6:
                return row
        return None

    def _as_number(self, value) -> float | None:
        """Convert worksheet values to float while ignoring bools and blanks."""
        if isinstance(value, bool) or value is None:
            return None
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return None
            try:
                return float(text)
            except ValueError:
                return None
        return None

    def _copy_or_create(self, output_name: str, sheet_name: str) -> Path:
        """Copy a template to output, or create a minimal workbook if missing."""
        output = self.output_dir / output_name
        template = self._template_path(output_name)
        if template.exists():
            copy2(template, output)
            return output
        if output.exists():
            return output

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        wb.save(output)
        return output

    def _template_path(self, output_name: str) -> Path:
        """Match either an exact template name or a -standard template name."""
        template = self.template_dir / output_name
        if template.exists():
            return template

        output_path = Path(output_name)
        standard_name = f"{output_path.stem}-standard{output_path.suffix}"
        return self.template_dir / standard_name

    def _clear_rows(self, ws: Worksheet, start_row: int, max_col: int) -> None:
        """Clear contiguous columns from start_row down."""
        for row in range(start_row, ws.max_row + 1):
            for col in range(1, max_col + 1):
                ws.cell(row=row, column=col).value = None

    def _clear_data_columns(self, ws: Worksheet, start_row: int, columns: tuple[int, ...]) -> None:
        """Clear only data columns in process workbooks."""
        for row in range(start_row, ws.max_row + 1):
            for col in columns:
                ws.cell(row=row, column=col).value = None

    def _mark_cells_red(self, ws: Worksheet, row: int, columns: tuple[int, ...]) -> None:
        """Mark the best point cells in red."""
        for col in columns:
            ws.cell(row=row, column=col).font = RED_FONT

    def describe_scan(self, result: CalibrationResult) -> str:
        """Return a short scan summary for UI messages or future logs."""
        config = result.config
        return (
            f"Feed{config.target_feed_id} "
            f"{format_phase(config.phase_start_deg)}to{format_phase(config.phase_end_deg)} "
            f"step {format_phase(config.phase_step_deg)}"
        )
